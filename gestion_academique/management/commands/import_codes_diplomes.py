# gestion_academique/management/commands/import_codes_diplomes.py
# -*- coding: utf-8 -*-
"""
Commande: import_codes_diplomes
Objectif: Mettre à jour Diplome.code à partir d'un Excel (.xlsx) via une clé composite stricte:
- matricule
- nom
- prenom
- date_de_naissance (date)
- lieu_de_naissance
- diplome
- niveau
- annee_academique

Le fichier Excel s'appelle 'diplomes.xlsx' et se trouve à la racine (même niveau que manage.py).

Usage exemples:
  Simulation (aucune écriture):
    python manage.py import_codes_diplomes --file ./diplomes.xlsx --dry-run

  Exécution réelle + rapport CSV:
    python manage.py import_codes_diplomes --file ./diplomes.xlsx --report ./rapport_final.csv

Prérequis: pip install openpyxl
"""
from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date
from typing import Any, Dict, List, Tuple, Set

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise CommandError("openpyxl requis. Installez-le: pip install openpyxl") from exc

from gestion_academique.models import Diplome


# ------------------ Helpers normalisation ------------------ #
def _strip_accents(s: str) -> str:
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_text(s: Any) -> str:
    """Normalise un texte: trim, str, déaccentue, compact espaces, lower."""
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    s = _strip_accents(s)
    s = " ".join(s.split())
    return s.lower()


def norm_code(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().upper()


def parse_excel_date(value: Any) -> date | None:
    """
    Convertit une date Excel en date Python.
    Accepte: cellule date, yyyy-mm-dd, dd/mm/yyyy, dd-mm-yyyy, yyyy/mm/dd, dd.mm.yyyy, ISO.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


@dataclass
class RowItem:
    key: str               # clé composite normalisée
    target_code: str       # code à appliquer
    raw: Dict[str, Any]    # valeurs brutes utiles au rapport


class Command(BaseCommand):
    help = "Met à jour 'Diplome.code' depuis un Excel via une clé composite stricte."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Chemin du fichier Excel (.xlsx)")
        parser.add_argument("--sheet", default=None, help="Nom de l'onglet (défaut: premier onglet)")
        # noms de colonnes attendues (personnalisables selon votre Excel)
        parser.add_argument("--col-matricule", default="matricule")
        parser.add_argument("--col-nom", default="nom")
        parser.add_argument("--col-prenom", default="prenom")
        parser.add_argument("--col-date-naissance", default="date_de_naissance")
        parser.add_argument("--col-lieu-naissance", default="lieu_de_naissance")
        parser.add_argument("--col-diplome", default="diplome")
        parser.add_argument("--col-niveau", default="niveau")
        parser.add_argument("--col-annee-academique", default="annee_academique")
        parser.add_argument("--col-code", default="code", help="Colonne contenant le code local à appliquer")
        # options
        parser.add_argument("--dry-run", action="store_true", help="Simulation sans écriture DB")
        parser.add_argument("--report", default=None, help="Chemin d'un CSV de rapport à produire (optionnel)")

    # --------------- Lecture Excel ---------------- #
    def _load_sheet(self, file_path: str, sheet_name: str | None):
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Feuille Excel vide.")
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]
        return headers, data_rows

    def _idx(self, headers: List[str], name: str) -> int:
        lowered = [h.lower() for h in headers]
        try:
            return lowered.index(name.lower())
        except ValueError:
            raise CommandError(f"Colonne '{name}' introuvable. En-têtes disponibles: {headers}")

    def _build_key_from_excel(self, headers: List[str], row: List[Any], colnames: Dict[str, str]) -> Tuple[str, Dict[str, Any]]:
        """
        Construit la clé composite normalisée depuis la ligne Excel + capture un dict brut
        pour le rapport.
        """
        ix = {k: self._idx(headers, v) for k, v in colnames.items()}

        matricule = norm_text(row[ix["matricule"]]) if ix["matricule"] < len(row) else ""
        nom = norm_text(row[ix["nom"]]) if ix["nom"] < len(row) else ""
        prenom = norm_text(row[ix["prenom"]]) if ix["prenom"] < len(row) else ""
        dob = parse_excel_date(row[ix["dob"]]) if ix["dob"] < len(row) else None
        dob_key = dob.isoformat() if dob else ""
        lieu = norm_text(row[ix["lieu"]]) if ix["lieu"] < len(row) else ""
        dipl = norm_text(row[ix["diplome"]]) if ix["diplome"] < len(row) else ""
        niveau = norm_text(row[ix["niveau"]]) if ix["niveau"] < len(row) else ""
        annee = norm_text(row[ix["annee"]]) if ix["annee"] < len(row) else ""

        key = "|".join([matricule, nom, prenom, dob_key, lieu, dipl, niveau, annee])
        raw = {
            "matricule": matricule, "nom": nom, "prenom": prenom,
            "date_de_naissance": dob_key, "lieu_de_naissance": lieu,
            "diplome": dipl, "niveau": niveau, "annee_academique": annee
        }
        return key, raw

    def _build_key_from_model(self, d: Diplome) -> str:
        matricule = norm_text(d.matricule)
        nom = norm_text(d.nom)
        prenom = norm_text(d.prenom)
        dob_key = d.date_de_naissance.isoformat() if d.date_de_naissance else ""
        lieu = norm_text(d.lieu_de_naissance)
        dipl = norm_text(d.diplome)
        niveau = norm_text(d.niveau)
        annee = norm_text(d.annee_academique)
        return "|".join([matricule, nom, prenom, dob_key, lieu, dipl, niveau, annee])

    # --------------- Conflits de codes --------------- #
    def _check_code_conflicts(self, target_codes: Set[str], planned_updates: Dict[int, str]) -> List[str]:
        """
        Vérifie si des codes cibles sont déjà utilisés par d'autres Diplome que ceux qu'on met à jour.
        planned_updates: map id_diplome -> nouveau_code
        """
        msgs: List[str] = []
        if not target_codes:
            return msgs
        for d in Diplome.objects.filter(code__in=target_codes).only("id", "code"):
            new_for_this = planned_updates.get(d.id)
            if new_for_this and new_for_this == (d.code or "").upper():
                continue  # même enregistrement, même code => ok
            msgs.append(f"Conflit: Diplome id={d.id} possède déjà le code '{d.code}'")
        return msgs

    # --------------- Rapport CSV --------------- #
    def _write_report(self, path: str, updated: List[Dict[str, Any]], unchanged: List[Dict[str, Any]], missing: List[Dict[str, Any]]):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["section", "matricule", "nom", "prenom", "date_de_naissance", "lieu_de_naissance",
                        "diplome", "niveau", "annee_academique", "old_code", "new_code", "db_id"])
            for row in updated:
                w.writerow(["UPDATED", row.get("matricule",""), row.get("nom",""), row.get("prenom",""),
                            row.get("date_de_naissance",""), row.get("lieu_de_naissance",""), row.get("diplome",""),
                            row.get("niveau",""), row.get("annee_academique",""), row.get("old_code",""),
                            row.get("new_code",""), row.get("db_id","")])
            for row in unchanged:
                w.writerow(["UNCHANGED", row.get("matricule",""), row.get("nom",""), row.get("prenom",""),
                            row.get("date_de_naissance",""), row.get("lieu_de_naissance",""), row.get("diplome",""),
                            row.get("niveau",""), row.get("annee_academique",""), row.get("old_code",""),
                            row.get("old_code",""), row.get("db_id","")])
            for row in missing:
                w.writerow(["MISSING", row.get("matricule",""), row.get("nom",""), row.get("prenom",""),
                            row.get("date_de_naissance",""), row.get("lieu_de_naissance",""), row.get("diplome",""),
                            row.get("niveau",""), row.get("annee_academique",""), "", row.get("new_code",""), ""])

    # --------------- Handle --------------- #
    def handle(self, *args, **opt):
        file_path: str = opt["file"]
        sheet_name: str | None = opt["sheet"]
        dry_run: bool = opt["dry_run"]
        report_path: str | None = opt["report"]

        headers, data_rows = self._load_sheet(file_path, sheet_name)

        # mapping de colonnes
        colnames = {
            "matricule": opt["col_matricule"],
            "nom": opt["col_nom"],
            "prenom": opt["col_prenom"],
            "dob": opt["col_date_naissance"],
            "lieu": opt["col_lieu_naissance"],
            "diplome": opt["col_diplome"],
            "niveau": opt["col_niveau"],
            "annee": opt["col_annee_academique"],
            "code": opt["col_code"],
        }

        # indices validation (lève si manquant)
        _ = {k: self._idx(headers, v) for k, v in colnames.items()}

        # Construire les items depuis Excel
        items: List[RowItem] = []
        for r in data_rows:
            key, raw = self._build_key_from_excel(
                headers, r,
                {
                    "matricule": colnames["matricule"],
                    "nom": colnames["nom"],
                    "prenom": colnames["prenom"],
                    "dob": colnames["dob"],
                    "lieu": colnames["lieu"],
                    "diplome": colnames["diplome"],
                    "niveau": colnames["niveau"],
                    "annee": colnames["annee"],
                }
            )
            code_idx = self._idx(headers, colnames["code"])
            target_code = norm_code(r[code_idx]) if code_idx < len(r) else ""
            if not key or not target_code:
                continue
            items.append(RowItem(key=key, target_code=target_code, raw=raw))

        if not items:
            raise CommandError("Aucune ligne exploitable (clé composite + code).")

        # Index DB par clé composite
        db_index: Dict[str, Diplome] = {}
        for d in Diplome.objects.all().only(
            "id", "matricule", "nom", "prenom", "date_de_naissance",
            "lieu_de_naissance", "diplome", "niveau", "annee_academique", "code"
        ):
            key = self._build_key_from_model(d)
            db_index[key] = d

        # Préparation des updates
        to_update: List[Diplome] = []
        unchanged_rows: List[Dict[str, Any]] = []
        missing_rows: List[Dict[str, Any]] = []
        planned_map: Dict[int, str] = {}  # id -> new_code (pour vérifier conflits)
        target_codes: Set[str] = set()

        for it in items:
            d = db_index.get(it.key)
            if not d:
                missing_rows.append({**it.raw, "new_code": it.target_code})
                continue
            new_code = it.target_code
            old_code = norm_code(d.code or "")
            if old_code == new_code:
                unchanged_rows.append({**it.raw, "old_code": old_code, "db_id": d.id})
                continue
            d.code = new_code
            to_update.append(d)
            planned_map[d.id] = new_code
            target_codes.add(new_code)

        # Conflits de codes avec d’autres enregistrements
        conflict_msgs = self._check_code_conflicts(target_codes, planned_map)
        if conflict_msgs:
            for m in conflict_msgs[:20]:
                self.stderr.write("❌ " + m)
            raise CommandError("Arrêt par sécurité: codes cibles déjà utilisés par d'autres Diplome.")

        self.stdout.write(
            f"Préparation: {len(items)} lignes Excel | updates={len(to_update)} | inchangés={len(unchanged_rows)} | introuvables={len(missing_rows)}"
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("Mode --dry-run: aucune écriture n'a été effectuée."))
        else:
            with transaction.atomic():
                for d in to_update:
                    d.save(update_fields=["code"])
            self.stdout.write(self.style.SUCCESS(f"✅ Terminé: {len(to_update)} codes mis à jour."))

        if report_path:
            updated_rows = [
                {
                    **{
                        "matricule": norm_text(d.matricule),
                        "nom": norm_text(d.nom),
                        "prenom": norm_text(d.prenom),
                        "date_de_naissance": d.date_de_naissance.isoformat() if d.date_de_naissance else "",
                        "lieu_de_naissance": norm_text(d.lieu_de_naissance),
                        "diplome": norm_text(d.diplome),
                        "niveau": norm_text(d.niveau),
                        "annee_academique": norm_text(d.annee_academique),
                    },
                    "old_code": "",  # Option: capturer l'ancien code avant l'affectation si vous le souhaitez
                    "new_code": norm_code(d.code),
                    "db_id": d.id,
                }
                for d in to_update
            ]
            self._write_report(report_path, updated_rows, unchanged_rows, missing_rows)
            self.stdout.write(self.style.SUCCESS(f"📄 Rapport écrit: {report_path}"))
